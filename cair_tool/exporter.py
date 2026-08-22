import csv
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List
 
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
 
COLUMN_GROUPS = {
    "A": [
        "comment_id",
        "comment_permalink",
        "comment_text",
        "parent_comment_id",
        "thread_level",
        "like_count",
        "reply_count",
        "published_at",
        "updated_at",
        "author_display_name",
        "author_channel_id",
        "author_channel_url",
        "video_id",
        "video_title",
        "video_url",
        "channel_name",
        "channel_id",
        "video_published_at",
        "video_view_count",
        "video_comment_count",
    ],
    "B": [
        "islamophobic",
        "islamophobia_category",
        "severity",
        "target",
        "triggering_span",
        "is_counterspeech",
        "overall_sentiment",
        "detected_language",
        "language_confidence",
        "english_translation",
        "model_confidence",
        "model_rationale",
        "model_version",
        "rubric_version",
    ],
    "C": [
        "human_reviewed",
        "reviewer_id",
        "review_date",
        "model_human_agreement",
        "final_label",
        "review_notes",
        "effective_label",
    ],
    "D": [
        "outlet_tier",
        "coverage_wave",
        "format",
    ],
    "E": [
        "collected_at",
        "collection_method",
        "sample_method",
        "comment_live_at_collection",
        "escalation_flag",
    ],
}
 
HEADER_ROW = [col for group in COLUMN_GROUPS.values() for col in group]
 
GROUP_COLORS = {
    "A": "FFF2CC",
    "B": "D9EAD3",
    "C": "C9DAF8",
    "D": "EAD1DC",
    "E": "FCE5CD",
}
 
 
def _group_for_header(header: str) -> str:
    for group, cols in COLUMN_GROUPS.items():
        if header in cols:
            return group
    return "A"
 
 
def _column_letter(header: str) -> str:
    return get_column_letter(HEADER_ROW.index(header) + 1)
 
 
def _build_comments_sheet(wb: Workbook, rows: List[Dict[str, object]]) -> None:
    ws = wb.create_sheet("comments")
    # Header row (written once).
    for i, header in enumerate(HEADER_ROW, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        color = GROUP_COLORS[_group_for_header(header)]
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(header) + 2, 12), 40)
 
    final_label_letter = _column_letter("final_label")
    islamophobic_letter = _column_letter("islamophobic")
    effective_label_col = HEADER_ROW.index("effective_label") + 1
 
    for row_index, row in enumerate(rows, start=2):
        for col_index, header in enumerate(HEADER_ROW, start=1):
            if col_index == effective_label_col:
                continue  # effective_label is a formula, written below
            ws.cell(row=row_index, column=col_index, value=row.get(header, None))
        ws.cell(
            row=row_index,
            column=effective_label_col,
            value=(
                f'=IF(AND(NOT(ISBLANK({final_label_letter}{row_index})), '
                f'{final_label_letter}{row_index}<>""), '
                f'{final_label_letter}{row_index}, {islamophobic_letter}{row_index})'
            ),
        )
    ws.freeze_panes = "A2"
 
 
def _build_by_video_sheet(wb: Workbook, video_rows: List[Dict[str, object]], comment_count: int) -> None:
    ws = wb.create_sheet("by_video")
    headers = [
        "video_id",
        "video_title",
        "video_url",
        "outlet_tier",
        "coverage_wave",
        "format",
        "video_view_count",
        "video_comment_count",
        "comments_collected",
        "yes_count",
        "not_sure_count",
        "counterspeech_count",
        "islamophobic_rate",
    ]
    ws.append(headers)
    for h in headers:
        ws.cell(row=1, column=headers.index(h) + 1, value=h).font = Font(bold=True)
    comment_video_col = _column_letter("video_id")
    # rates key off effective_label so human overrides flow through
    comment_label_col = _column_letter("effective_label")
    comment_counterspeech_col = _column_letter("is_counterspeech")
    for row_idx, video in enumerate(video_rows, start=2):
        ws.cell(row=row_idx, column=1, value=video["video_id"])
        ws.cell(row=row_idx, column=2, value=video["video_title"])
        ws.cell(row=row_idx, column=3, value=video["video_url"])
        ws.cell(row=row_idx, column=4, value=video["outlet_tier"])
        ws.cell(row=row_idx, column=5, value=video["coverage_wave"])
        ws.cell(row=row_idx, column=6, value=video["format"])
        ws.cell(row=row_idx, column=7, value=video["video_view_count"])
        ws.cell(row=row_idx, column=8, value=video["video_comment_count"])
        id_cell = ws.cell(row=row_idx, column=1).coordinate
        ws.cell(row=row_idx, column=9, value=f"=COUNTIFS(comments!{comment_video_col}:{comment_video_col},{id_cell})")
        ws.cell(row=row_idx, column=10, value=f'=COUNTIFS(comments!{comment_video_col}:{comment_video_col},{id_cell},comments!{comment_label_col}:{comment_label_col},"yes")')
        ws.cell(row=row_idx, column=11, value=f'=COUNTIFS(comments!{comment_video_col}:{comment_video_col},{id_cell},comments!{comment_label_col}:{comment_label_col},"not_sure")')
        ws.cell(row=row_idx, column=12, value=f"=COUNTIFS(comments!{comment_video_col}:{comment_video_col},{id_cell},comments!{comment_counterspeech_col}:{comment_counterspeech_col},1)")
        # rate = yes_count (col J) / comments_collected (col I)
        ws.cell(row=row_idx, column=13, value=f"=IFERROR(J{row_idx}/I{row_idx},0)")
        ws.cell(row=row_idx, column=13).number_format = "0.0%"
 
 
def _formula_block(ws, start_row: int, header: str, field_col: str, group_values: List[str], label_col: str) -> int:
    ws.cell(row=start_row, column=1, value=header).font = Font(bold=True)
    ws.cell(row=start_row + 1, column=1, value="group").font = Font(bold=True)
    ws.cell(row=start_row + 1, column=2, value="K_yes").font = Font(bold=True)
    ws.cell(row=start_row + 1, column=3, value="N_total").font = Font(bold=True)
    ws.cell(row=start_row + 1, column=4, value="rate").font = Font(bold=True)
    for idx, value in enumerate(group_values, start=0):
        row = start_row + 1 + idx + 1
        ws.cell(row=row, column=1, value=value)
        ws.cell(row=row, column=2, value=f'=COUNTIFS(comments!{field_col}:{field_col},"{value}",comments!{label_col}:{label_col},"yes")')
        ws.cell(row=row, column=3, value=f'=COUNTIFS(comments!{field_col}:{field_col},"{value}")')
        ws.cell(row=row, column=4, value=f"=IFERROR(B{row}/C{row},0)")
        ws.cell(row=row, column=4).number_format = "0.0%"
    all_row = start_row + len(group_values) + 2
    ws.cell(row=all_row, column=1, value="ALL").font = Font(bold=True)
    ws.cell(row=all_row, column=2, value=f'=COUNTIFS(comments!{label_col}:{label_col},"yes")')
    ws.cell(row=all_row, column=3, value=f"=COUNTA(comments!{label_col}:{label_col})-1")
    ws.cell(row=all_row, column=4, value=f"=IFERROR(B{all_row}/C{all_row},0)")
    ws.cell(row=all_row, column=4).number_format = "0.0%"
    return all_row + 2
 
 
def _build_rates_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("rates")
    label_col = _column_letter("effective_label")
    tier_col = _column_letter("outlet_tier")
    wave_col = _column_letter("coverage_wave")
    format_col = _column_letter("format")
    next_row = 1
    next_row = _formula_block(ws, next_row, "By outlet_tier", tier_col, ["national", "local_san_diego", "independent_commentary"], label_col)
    next_row = _formula_block(ws, next_row, "By coverage_wave", wave_col, ["breaking_news", "manifesto_radicalization", "official_briefings", "community_response"], label_col)
    _formula_block(ws, next_row, "By format", format_col, ["short", "long_form"], label_col)
 
 
def _build_unique_accounts_sheet(wb: Workbook, rows: List[Dict[str, object]]) -> None:
    ws = wb.create_sheet("unique_accounts")
    ws.append(["metric", "value"])
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)
    # Compute in Python to avoid array-formula div-by-zero on blank cells.
    counts = Counter([(row.get("author_channel_id") or "(unknown)") for row in rows])
    ws.cell(row=2, column=1, value="distinct_author_channel_id_count")
    ws.cell(row=2, column=2, value=len(counts))
    ws.cell(row=3, column=1, value="total_comments")
    ws.cell(row=3, column=2, value=len(rows))
    ws.append(["top_commenter_channel_id", "comment_count"])
    ws.cell(row=4, column=1).font = Font(bold=True)
    ws.cell(row=4, column=2).font = Font(bold=True)
    for idx, (channel_id, count) in enumerate(counts.most_common(10), start=5):
        ws.cell(row=idx, column=1, value=channel_id)
        ws.cell(row=idx, column=2, value=count)
    note_row = 5 + min(len(counts), 10) + 1
    ws.cell(row=note_row, column=1, value="A few high-frequency accounts can indicate coordinated rather than organic activity.").font = Font(italic=True, size=9)
 
 
def _build_validation_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("validation")
    ws.append(["metric", "formula"])
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)
    reviewed_col = _column_letter("human_reviewed")
    agreement_col = _column_letter("model_human_agreement")
    ws.append(["reviewed_count", f"=COUNTIFS(comments!{reviewed_col}:{reviewed_col},1)"])
    ws.append(["agreement_count", f"=COUNTIFS(comments!{reviewed_col}:{reviewed_col},1,comments!{agreement_col}:{agreement_col},1)"])
    ws.append(["agreement_rate", "=IFERROR(B3/B2,0)"])
    ws.cell(row=4, column=2).number_format = "0.0%"
    ws.append([])
    ws.append(["A low reviewed_count means the rate is indicative, not validated."])
    ws.cell(row=6, column=1).font = Font(italic=True, size=9)
 
 
def _build_escalations_sheet(wb: Workbook, rows: List[Dict[str, object]]) -> None:
    ws = wb.create_sheet("escalations")
    headers = ["comment_id", "comment_permalink", "comment_text", "english_translation", "islamophobia_category", "severity", "escalation_flag"]
    ws.append(headers)
    for h in headers:
        ws.cell(row=1, column=headers.index(h) + 1).font = Font(bold=True)
    flagged = False
    for row in rows:
        if row.get("escalation_flag"):
            ws.append([row.get(header) for header in headers])
            flagged = True
    if not flagged:
        ws.append(["(none flagged in this run)"])
    ws.append([])
    ws.append(["Human triage required; do not auto-report. Refer to YouTube/FBI (1-800-CALL-FBI) if appropriate."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9)
 
 
def _build_methodology_sheet(wb: Workbook, video_count: int, comment_count: int, date_range: str, config: Dict[str, object], sample_method: str) -> None:
    ws = wb.create_sheet("methodology")
    ws.append(["field", "value"])
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)
    quota = config.get("daily_quota_ceiling", config.get("daily_quota_limit", ""))
    ws.append(["comment_count", comment_count])
    ws.append(["video_count", video_count])
    ws.append(["date_range", date_range])
    ws.append(["collection_method", "YouTube Data API v3"])
    ws.append(["sample_method", sample_method])
    ws.append(["model_version", config.get("model")])
    ws.append(["rubric_version", config.get("rubric_version")])
    ws.append(["quota_ceiling", quota])
    ws.append(["scope_caveat", "This dataset describes the YouTube comment conversation on this coverage, not public sentiment generally. Comment sections skew opinionated, and rates are reported with denominators."])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    ws.cell(row=10, column=2).alignment = Alignment(wrap_text=True, vertical="top")
 
 
def export_workbook(rows: List[Dict[str, object]], video_rows: List[Dict[str, object]], config: Dict[str, object]) -> Path:
    out_dir = Path(config["output_dir"])
    out_dir.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _build_comments_sheet(workbook, rows)
    _build_by_video_sheet(workbook, video_rows, len(rows))
    _build_rates_sheet(workbook)
    _build_unique_accounts_sheet(workbook, rows)
    _build_validation_sheet(workbook)
    _build_escalations_sheet(workbook, rows)
    date_range = _build_date_range(rows)
    _build_methodology_sheet(workbook, len({row.get("video_id") for row in rows}), len(rows), date_range, config, config.get("sample_method", ""))
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    output_path = out_dir / f"comments_export_{timestamp}.xlsx"
    workbook.save(output_path)
    csv_path = out_dir / f"comments_export_{timestamp}.csv"
    _export_csv(rows, csv_path)
    return output_path
 
 
def _export_csv(rows: List[Dict[str, object]], csv_path: Path) -> None:
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=HEADER_ROW)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in HEADER_ROW})
 
 
def _build_date_range(rows: List[Dict[str, object]]) -> str:
    dates = [row.get("published_at") for row in rows if row.get("published_at")]
    try:
        sorted_dates = sorted(dates)
        if sorted_dates:
            return f"{sorted_dates[0]} to {sorted_dates[-1]}"
    except Exception:
        pass
    return "unknown"
